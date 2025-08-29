# Format an Excel file with openpyxl

import openpyxl
from openpyxl.styles import Font, PatternFill

#Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales"

# Insert data
data = [("Product", "Quantity", "Price"),
        ("Mouse", 20, 15),
        ("Keyboard", 35, 30),
        ("Monitor", 10, 200)]

for row in data:
    ws.append(row)
    
# Style the header row
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    
wb.save("Sales_Report.xlsx")
    
    